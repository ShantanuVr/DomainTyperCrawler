from selenium import webdriver
from BeautifulSoup import BeautifulSoup
from urlparse import urlparse
import pandas as pd
import openpyxl as op
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
import re

url = "https://domaintyper.com/top-websites"
xl_name = "Crawler.xlsx"
sheet_name = "URLs"

def lookupPrimary(url, driver, crawled_urls):
	driver.get(url)
	html = driver.page_source.encode("utf-8")
	soup = BeautifulSoup(html)
	urls = soup.findAll("a")
	urls_list = list()
	for a in urls:
		if "top-websites" in a.get("href"):
			if (a.get("href")) and (a.get("href") not in urls_list):
				urls_list.append("https://domaintyper.com" + a.get("href"))
	for b in urls_list:
		lookupSecondary(b, driver, crawled_urls)
	
def lookupSecondary(url, driver, crawled_urls):
	driver.get(url)
	html = driver.page_source.encode("utf-8")
	soup = BeautifulSoup(html)
	# columns = list()
	rows = list()
	# columns = soup.findAll('div', attrs={"class" : "pagingDivDisabled"})
	rows = soup.findAll('div', attrs={"class" : "pagingDiv"})
	urlTable = soup.findAll('a', attrs={"title" : re.compile('^(Whois Record)')})

	if any(rows) and "Next" in rows[-1]:
		crawl_urls(urlTable, crawled_urls)
		nextPageURL = soup.findAll('a', attrs={"class" : "pagingLink"})
		nextPageURL = "https://domaintyper.com" + nextPageURL[-1].get("href")
		
		lookupSecondary(nextPageURL, driver, crawled_urls)
	
def crawl_urls(urlTable, crawled_urls):
	for a1 in urlTable:
		for a2 in a1.contents:
			if a2 not in crawled_urls:
				crawled_urls.append(a2)
				
def load_to_excel(lst):
	"""
	Load the list into excel file using pandas
	"""
	# Load list to dataframe
	df = pd.DataFrame(lst)
	df.index += 1  # So that the excel column starts from 1
	
	# Write dataframe to excel
	xlw = pd.ExcelWriter(xl_name)
	df.to_excel(xlw, sheet_name=sheet_name, index_label="#", header=["URL"])
	xlw.save()


def format_excel(xl, sheet="Sheet1"):
	"""
	Get the excel file path and format the file
	If no sheet name is passed, by default take Sheet1
	"""
	# Open the excel file
	wb = op.load_workbook(xl)
	ws = wb.get_sheet_by_name(sheet)
	
	# Freeze panes
	ws.freeze_panes = "B2"
	
	# Adjust column width
	cols = ("A", "B")
	widths = (5, 80)
	
	for combo in zip(cols, widths):
		ws.column_dimensions[combo[0]].width = combo[1]
	
	# define color formmatting
	blue_fill = op.styles.PatternFill(start_color="00aadd",
									fill_type='solid')
	
	# define border style
	thin_border = op.styles.borders.Border(left=op.styles.Side(style='thin'),
										right=op.styles.Side(style='thin'),
										top=op.styles.Side(style='thin'),
										bottom=op.styles.Side(style='thin'))
	
	# define Text wrap
	text_wrap = op.styles.Alignment(wrap_text=True)
	
	# Format the header row
	for row in range(1, 2):  # Loop only the 1st row
		for col in range(1, ws.max_column + 1):  # loop through all columns
			ws.cell(row=row, column=col).fill = blue_fill
	
	# Format all cells
	for row in ws.iter_rows():
		for cell in row:
			# Draw borders
			cell.border = thin_border
	
			# Wrap all columns
			cell.alignment = text_wrap
	
	# Save back as same file name
	wb.save(xl)
		
if __name__ == "__main__":
	binary = FirefoxBinary(r"C:\Users\Shantanu\AppData\Local\Mozilla Firefox\firefox.exe")
	driver = webdriver.Firefox(firefox_binary=binary)
	crawled_urls = list()
	lookupPrimary(url, driver, crawled_urls)
	driver.quit()
	# print "FULL URLs LIST"
	# print len(set(url_list))
	
	print "CRAWLED URLs LIST"
	print len(set(crawled_urls))
	
	# Load the match list to excel
	load_to_excel(crawled_urls)
	
	# Format the excel file
	format_excel(xl_name, sheet_name)